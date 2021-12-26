#! /usr/bin/env python3
# ! -*- coding: utf-8 -*-
# yaozhengyi.ah@chinatelecom Anhui Usage scence
# 5G SA UPF(SHOW L34FILTER; SHOW L34FILTERGROUP; SHOW L7FILTER;SHOW L7FILTERGROUP) to JT rule xlsx file
# (DONE) 20210801 20211219实现 根据定向流量规则表格做全量l34/l7规则更新
# (DONE) 20211225 增加定向流量规则表格l34规则判重
# (DOING) 20211219 全量更新脚本测试 中国电信集团生产任务[2021]612号 爱奇艺视频调度-20211118_new 测试OK!
# (TODO) 20211219 优化空间，读取现网UPF规则时，只对需要变动的RG进行加工，非全量RG
# (TODO) 20211225 输出重复项中存在IPv6地址在表格的格式与程序输出的格式不一致，查找表格不方便
# 常见RG
# 3532707700    阿里云
# 3532708000    VIVO
# 3532703600    阿里巴巴
# 3532704600    爱奇艺
# 3532707000    百度
# 1600000002    哔哩哔哩
# 3532707600    手淘天猫
# 3532707100    网易
# 3532704700    优酷
# 3532706700    今日头条
# 3532704500    腾讯视频
# 3532706100    天翼云盘

import csv
import os
import re
import time
import copy
from datetime import datetime
# 以下两个模块openpyxl和netaddr需要额外安装
import netaddr
import openpyxl
import warnings

warnings.simplefilter("ignore")


# (DONE) 20211225 定向流量规则表格l34规则判重,区分IPv4/IPv6 降低运算复杂度
def fncheckdup(inl34_vlaue, inl34_ipnetwork_v4, inl34_ipnetwork_v6):
    if inl34_vlaue.version == 4:
        if len(inl34_ipnetwork_v4) > 1:
            position1 = inl34_ipnetwork_v4.index(inl34_vlaue)
            for everysubnet1 in inl34_ipnetwork_v4[position1 + 1:]:
                s1 = netaddr.IPSet(inl34_vlaue)
                s2 = netaddr.IPSet(everysubnet1)
                if (s1 & s2).size > 0:
                    print(u'\033[0;31;40m!!!发现l34规则存在IPv4重复条目\033[0m', inl34_vlaue, everysubnet1)
                break
    else:
        # (TODO) 20211225 输出重复项中存在IPv6地址在表格的格式与程序输出的格式不一致，查找表格不方便
        if len(inl34_ipnetwork_v6) > 1:
            position2 = inl34_ipnetwork_v6.index(inl34_vlaue)
            for everysubnet2 in inl34_ipnetwork_v6[position2 + 1:]:
                s1 = netaddr.IPSet(inl34_vlaue)
                s2 = netaddr.IPSet(everysubnet2)
                if (s1 & s2).size > 0:
                    print(u'\033[0;31;40m!!!发现l34规则存在IPv6重复条目\033[0m', inl34_vlaue, everysubnet2)
                break


# 删除为空的输出文件
def fndelemptyoutfile(infile):
    # (20211224)先判断文件是否存在
    if os.path.exists(infile):
        sz = os.stat(infile).st_size
        if sz == 0:
            os.remove(infile)


# 获得三四层过滤规则组名(L34FILTERGROUP_INFO)与三四层过滤器名(L34FILTER_INFO)的对应关系
def fn_get_cur_l34filtergrp(inl34filtergrp):
    fnl34filtergrp_dic = {}
    fnl34filtergrp_all = []
    fnl34filterinfo_all = []
    with open(inl34filtergrp, 'r', encoding='utf-8') as infile:
        reader = csv.reader(infile)
        # 跳过首行
        next(reader)
        for line in reader:
            l34filtergrp = line[0]
            l34filterinfo = line[1]
            if l34filtergrp in fnl34filtergrp_dic:
                fnl34filtergrp_dic[l34filtergrp].append(l34filterinfo)
            else:
                fnl34filtergrp_all.append(l34filtergrp)
                fnl34filtergrp_dic[l34filtergrp] = [l34filterinfo]
            if l34filterinfo in fnl34filterinfo_all:
                print(u'!!!L34FILTERGROUP_INFO文件中发现重复规则，建议人工核查', l34filterinfo)
            fnl34filterinfo_all.append(l34filterinfo)
    return fnl34filtergrp_dic, fnl34filtergrp_all, fnl34filterinfo_all


# 获得三四层过滤规则组名(L7FILTERGROUP_INFO)与三四层过滤器名(L7FILTER_INFO)的对应关系
def fn_get_cur_l7filtergrp(inl7filtergrp):
    fnl7filtergrp_dic = {}
    fnl7filtergrp_all = []
    fnl7filterinfo_all = []
    with open(inl7filtergrp, 'r', encoding='utf-8') as infile:
        reader = csv.reader(infile)
        # 跳过首行
        next(reader)
        for line in reader:
            l7filtergrp = line[0]
            l7filterinfo = line[1]
            if l7filtergrp in fnl7filtergrp_dic:
                fnl7filtergrp_dic[l7filtergrp].append(l7filterinfo)
            else:
                fnl7filtergrp_all.append(l7filtergrp)
                fnl7filtergrp_dic[l7filtergrp] = [l7filterinfo]
            if l7filterinfo in fnl7filterinfo_all:
                print(u'!!!L7FILTERGROUP_INFO文件中发现重复规则，建议人工核查', l7filterinfo)
            fnl7filterinfo_all.append(l7filterinfo)
    return fnl7filtergrp_dic, fnl7filtergrp_all, fnl7filterinfo_all


def fn_get_rg_from_l34filtergrp(indictkey):
    reobj_rg_name = re.compile(r"_(\d{9,10})")
    match_rg_name = reobj_rg_name.search(indictkey)
    # 得到 过滤规则名称 中的RG值
    if match_rg_name:
        out_rg_name = match_rg_name.group(1)
    else:
        # print(u'L34FILTERGROUP没有找到9或10位长度的RG值', indictkey)
        return None
    return out_rg_name


def fn_get_cur_l34filter(inl34filter_info, in_l34filterinfo_all, chk_rgnum):
    # l34层规则判重开关， 1为开启判断，0为不开启判断
    flag_chk_dupl = 0
    # 输出两个list，分别为规则名称和5元组
    # 实际L34 filtername(过滤规则名称)命名 : l34_f_3532704700_ipv6_20210305_10, l34_f_3532704700_ipv6_169 cry~~~
    # 建议L34 filtername(过滤规则名称) 规范命名 : l34_f_RG_ipv4_YYYYMMDD_NUM/l34_f_RG_ipv6_YYYYMMDD_NUM NUM为四位数字
    # (DONE) 20210330 增加 过滤规则名称 与 过滤规则组的判断
    with open(inl34filter_info, 'r', encoding='utf-8') as infile:
        reader = csv.reader(infile)
        # 跳过首行
        next(reader)
        l34filternamelist = []
        l34filtersubnetlist = []
        duplnum = 0
        for line in reader:
            if "_f_any_" in line or "_f_dns_" in line:
                continue
            reobj_rg_name = re.compile(r"f_(\d{9,10})_")
            match_rg_name = reobj_rg_name.search(line[0])
            # 得到 过滤规则名称 中的RG值
            if match_rg_name:
                fnrg_name = match_rg_name.group(1)
            else:
                continue
                # print(u'L34FILTER', line[0], '没有找到9或10位长度的RG值')
            # 跳过默认规则 RG 3000000000
            if fnrg_name == '3000000000':
                continue
            # 为减少计算量，只对需要更新的RG值获取l34规则
            if not fnrg_name == chk_rgnum:
                continue
            if line[0] not in in_l34filterinfo_all:
                print(u'!!!过滤规则名称(L34FILTER)', line[0], '未找到对应的过滤规则组(L34FILTERGROUP)')
                continue
            if line[1] == 'IPV4':
                subnet = line[2].strip() + "/" + line[3].strip()
            elif line[1] == 'IPV6':
                subnet = line[5].strip() + "/" + line[6].strip()
            l34filternamelist.append(line[0].strip())
            l34filtersubnetlist.append(subnet)
            # l34规则判重开关,打开后增加计算量
            if flag_chk_dupl == 1:
                if len(l34filtersubnetlist) > 1:
                    for everycursubnet in l34filtersubnetlist[0:-1]:
                        k1 = netaddr.IPNetwork(everycursubnet)
                        k2 = netaddr.IPNetwork(subnet)
                        # (20210414) 只当同为IPv4或者IPv6时进行判重,利还是弊？
                        if k1.version == k2.version:
                            # if netaddr.IPNetwork(subnet) in curipset: //20210328 废弃 速度慢
                            # 优化检测速度，求交集比求子集准确，size比leng() 更快（实际测试速度从759秒提升到181秒）
                            if (netaddr.IPSet(k2) & netaddr.IPSet(k1)).size > 0:
                                elementindex = l34filtersubnetlist.index(everycursubnet)
                                duplnum += 1
                                print(u'!!!发现l34规则', line[0], subnet, '与l34规则存在重复地址',
                                      l34filternamelist[elementindex], everycursubnet)
                                break
    # 范例数据:
    # ['l34_f_3532704700_ipv4_1', 'l34_f_3532704700_ipv4_10', 'l34_f_3532704700_ipv4_100']
    # ['101.226.184.0/25', '111.13.133.0/24', '183.61.238.0/24', '116.211.221.12/32']
    if duplnum > 0:
        print(u'一共发现%d组重复的l34规则' % duplnum)
    return l34filternamelist, l34filtersubnetlist,


def fn_get_cur_l7filter(inl7filter_info, in_l7filterinfo_all, chk_rgnum):
    # l7层规则判重开关， 1为开启判断，0为不开启
    flag_chk_dupl = 0
    # 输出两个list，分别为规则名称,URL
    # 实际L7 filtername(过滤规则名称)命名 : l7_f_1000000001_ipv4_2, l7_f_3532703300_76,l7_f_3532706700_143 cry~~~
    # 建议L7 filtername(过滤规则名称) 规范命名 : l7_f_RG_YYYYMMDD_NUM RG为9-10位数字，YYYYMMDD为日期，NUM为四位数字
    # (DONE) 20210401 增加 过滤规则名称 与 过滤规则组的判断
    # (DONE) 20210401 增加 过滤规则URL重复判断
    # l7规则判重开关,打开后耗费时间较多
    with open(inl7filter_info, 'r', encoding='utf-8') as infile:
        reader = csv.reader(infile)
        # 跳过首行
        next(reader)
        l7filternamelist = []
        l7filterurllist = []
        duplnum = 0
        for line in reader:
            if "filter" in line:
                continue
            reobj_rg_name = re.compile(r"f_(\d{9,10})_")
            match_rg_name = reobj_rg_name.search(line[0])
            if match_rg_name:
                fnrg_name = match_rg_name.group(1)
            else:
                # print(u'L7FILTER', line[0], '没有找到9或10位长度的RG值')
                continue
            # 为减少计算量，只对需要更新的RG值获取l7规则
            if not fnrg_name == chk_rgnum:
                continue
            if line[0] not in in_l7filterinfo_all:
                print(u'!!!过滤规则名称(L7FILTER)', line[0], '未找到对应的过滤规则组(L7FILTERGROUP)')
                continue
            url = line[1].strip()
            l7filternamelist.append(line[0])
            l7filterurllist.append(url)
            if flag_chk_dupl == 1:
                if len(l7filterurllist) > 1:
                    for existurl in l7filterurllist[0:-1]:
                        if url == existurl:
                            existurlindex = l7filterurllist.index(existurl)
                            duplnum += 1
                            print(u'!!!发现l7规则', line[0], url, '与l7规则存在重复URL',
                                  l7filternamelist[existurlindex], existurl)
                            break
    # 范例数据:
    # ['l7_f_3532704700_10', 'l7_f_3532704700_11', 'l7_f_3532704700_12']
    # ['*.cibntv.net', '*.youku.com', '*.uczzd.com']
    if duplnum > 0:
        print(u'一共发现%d组重复的l7规则' % duplnum)
    return l7filternamelist, l7filterurllist


def fn_get_cur_dpi_xlsx(inxlsxfile):
    # l34层规则判重开关， 1为开启判断，0为不开启判断
    flag_chk_dupl = 1
    jtrulewb = openpyxl.load_workbook(inxlsxfile)
    jtrulesheet = jtrulewb[jtrulewb.sheetnames[0]]
    jtrulenumber = jtrulesheet.max_row
    # C3  定向流量RG值
    # I列，IPv4/v6地址
    # J列，掩码
    # N列，URL
    # P列，维护信息，新增/删除/修改/不变/生效中/已报备/本次新增/本次删除 不同规则文件此列填的五花八门
    # print(u'规则共' + str(jtrulenumber) + u'行')
    rg_txt = jtrulesheet["C3"].value
    productname = jtrulesheet["A3"].value
    # (DONE) 20211217 RG查找优化
    dpirg = '0'
    digital_all = re.findall(r'\d+(?!\d+)', rg_txt)
    for digital in digital_all:
        if len(digital) == 10:
            dpirg = digital
    if dpirg == '0':
        print(u'!!!定向浏览规则表格文件中的RG长度可能有误，建议人工复查')
    col_i = jtrulesheet["I"]
    # print(col_i[0].value,col_i[1].value,col_i[2].value,col_i[3].value,col_i[-3].value,col_i[-2].value)
    col_j = jtrulesheet["J"]
    col_n = jtrulesheet["N"]
    # print(col_n[0].value, col_n[1].value, col_n[2].value,col_n[3].value,col_n[-3].value，col_n[-2].value)
    col_p = jtrulesheet["P"]
    jtrulewb.close()
    # l34是IPNetwork组成的list
    l34_add = []
    l34_dele = []
    l7_add = []
    l7_dele = []
    l34_add_num = 0
    l7_add_num = 0
    l34_del_num = 0
    l7_del_num = 0
    # (20211224) 存量l34规则判断重复
    l34_ipnetwork_v4 = []
    l34_ipnetwork_v6 = []
    for row_num in range(2, jtrulenumber):
        l34ip_value = col_i[row_num].value
        l34mask_value = col_j[row_num].value
        l7host_value = col_n[row_num].value
        flag_value = col_p[row_num].value
        if l34ip_value is None:
            if l7host_value is not None:
                # (DONE) 20210906 对N列URL为https://或http://开头,剔除协议部分 fixed in 20210908
                if l7host_value.startswith('https://'):
                    l7host_value = l7host_value[8::]
                elif l7host_value.startswith('http://'):
                    l7host_value = l7host_value[7::]
                # (20211219) 只有有删除关键字就认为是不需要的
                if flag_value is not None:
                    if u'删除' in flag_value:
                        l7_del_num += 1
                        l7_dele.append(l7host_value.strip())
                    else:
                        l7_add_num += 1
                        l7_add.append(l7host_value.strip())
                else:
                    l7_add_num += 1
                    l7_add.append(l7host_value.strip())
        elif l34ip_value is not None:
            # （20210401) 检查l34地址是否是有效的地址
            try:
                # I列给出掩码，忽略J列
                if '/' in l34ip_value:
                    l34_vlaue = netaddr.IPNetwork(l34ip_value.strip())
                # I列未给出掩码,取J列作为掩码
                else:
                    l34_vlaue = netaddr.IPNetwork(l34ip_value.strip() + '/' + l34mask_value.strip())
            except Exception as exp:
                print(u'I' + str(row_num) + '单元格内容有误，请人工检查!', l34ip_value.strip(), l34mask_value.strip())
                print(u'错误类型为', exp)
                break
            l34_prefix = l34_vlaue.prefixlen
            if l34_vlaue.version == 4 and l34_prefix <= 21:
                print(u'\033[0;31;40mIPv4前缀小于20,请额外注意\033[0m', str(l34_vlaue))
            if l34_vlaue.version == 6 and l34_prefix < 64:
                print(u'\033[0;31;40mIPv6前缀小于64,请额外注意\033[0m', str(l34_vlaue))
            if flag_value is not None:
                if u'删除' in flag_value:
                    l34_del_num += 1
                    l34_dele.append(l34_vlaue)
                else:
                    l34_add_num += 1
                    l34_add.append(l34_vlaue)
                    if l34_vlaue.version == 4:
                        l34_ipnetwork_v4.append(l34_vlaue)
                    if l34_vlaue.version == 6:
                        l34_ipnetwork_v6.append(l34_vlaue)
            else:
                l34_add_num += 1
                l34_add.append(l34_vlaue)
                if l34_vlaue.version == 4:
                    l34_ipnetwork_v4.append(l34_vlaue)
                if l34_vlaue.version == 6:
                    l34_ipnetwork_v6.append(l34_vlaue)
    # 清洗现有l34规则，去除存在删除要求的l34规则,
    # (TODO) 优化空间：使用ipset求差集 l34_add-l34_dele 问题：与集团表格无法一致
    # print(l34_dele)
    l34_add = list(set(l34_add))
    l34_dele = list(set(l34_dele))
    l34_ipnetwork_v4 = list(set(l34_ipnetwork_v4))
    l34_ipnetwork_v6 = list(set(l34_ipnetwork_v6))
    l34_add.sort()
    l34_ipnetwork_v4.sort()
    l34_ipnetwork_v6.sort()
    # 20211225 输出既要求删除有存量的l34条目 后续可以加上时间的判断
    conflictlist = []
    l34_add_temp = copy.deepcopy(l34_add)
    for x1 in l34_add:
        if x1 in l34_dele:
            conflictlist.append(x1)
            l34_add_temp.remove(x1)
            if x1.version == 4:
                l34_ipnetwork_v4.remove(x1)
            else:
                l34_ipnetwork_v6.remove(x1)
    print(u'表格中共发现%d组重复的l34规则' % (len(conflictlist)))
    for every_conflict in conflictlist:
        print(every_conflict)
    l34_add = l34_add_temp
    # (20211225) 新规则进行l34判重
    if flag_chk_dupl == 1:
        for x2 in l34_add:
            fncheckdup(x2, l34_ipnetwork_v4, l34_ipnetwork_v6)
    print(u'规则共' + str(jtrulenumber) + u'行')
    print(u'\033[0;30;42m' + productname, 'RG', dpirg + u'\033[0m')
    print(u'三层IP规则(非删除/删除) ' + str(l34_add_num) + '/' + str(l34_del_num) + ' 七层URL规则(非删除/删除) ' + str(
        l7_add_num) + '/' + str(l7_del_num) + u'\033[0m')
    print(u'三层IP规则新增 ' + str(len(l34_add)) + ' 七层URL规则新增 ' + str(len(l7_add)) + u'\033[0m')
    return dpirg, productname, l34_add, l7_add


def fn_compare_rough(rgnum, in_old_l34_name, in_old_l7_name, in_new_l34, in_new_l7, in_upfnum):
    configday = str(datetime.now().strftime('%Y%m%d'))
    config_time = str(datetime.now().strftime('%Y%m%d%H%M%S'))
    outdir = "allrules/" + config_time + "_RG_" + rgnum + "_" + in_upfnum
    # 建议规范写法
    # l34filername:l34_f_{RG}_{YYYYMMDD}_{NUM}
    # l7filername:l7_f_{RG}_{YYYYMMDD}_{NUM}
    # l34filtergrp:l34_g_{RG}_1
    # l7filtergrp:l7_g_{RG}_1
    # RG：9-10位长度的整数
    # YYYYMMDD: 8位长度年月日
    # (20211219) 由于l34规则存在较多，从3位扩充到4位长度
    # NUM：四位长度整数
    if not os.path.exists(outdir):
        os.mkdir(outdir)
    # l34 输出配置命令
    l34_fl_add = outdir + '/l34_f_add.txt'
    l34_flgp_add = outdir + '/l34_g_add.txt'
    l34_fl_del = outdir + '/l34_f_del.txt'
    l34_flgp_del = outdir + '/l34_g_del.txt'
    # l34 配置模板
    tmpl_l34_f_v4_add = 'ADD L34FILTER:FILTERNAME="l34_f_%s_%s_%04d",IPTYPE="IPV4",IPV4SERVERIP="%s",' \
                        'IPV4SERVERIPMASK=%s,PROTOCOL="ANY",SERVERPORTSTART=0,SERVERPORTEND=0; '
    tmpl_l34_f_v6_add = 'ADD L34FILTER:FILTERNAME="l34_f_%s_%s_%04d",IPTYPE="IPV6",IPV6SERVERIP="%s",' \
                        'IPV6SERVERIPMASK=%s,PROTOCOL="ANY",SERVERPORTSTART=0,SERVERPORTEND=0;'
    tmpl_l34_f_del = 'DEL L34FILTER:FILTERNAME="%s" '
    tmpl_l34_g_add = 'ADD L34FILTERGROUP:GROUPNAME="l34_g_%s_1",L34FILTERNAME="l34_f_%s_%s_%04d" '
    tmpl_l34_g_del = 'DEL L34FILTERGROUP:GROUPNAME="l34_g_%s_1",L34FILTERNAME="%s" '
    l34numadd = 0
    l34numdel = 0
    if len(in_new_l34) > 0:
        with open(l34_fl_add, 'w', encoding='utf-8') as f_l34_f_add, \
                open(l34_flgp_add, 'w', encoding='utf-8') as f_l34_g_add:
            for every_l34 in in_new_l34:
                ip, prefix = str(every_l34).split('/')
                ipver = every_l34.version
                l34numadd += 1
                # l34规则是ipv4场景
                if ipver == 4:
                    f_l34_f_add.write(tmpl_l34_f_v4_add % (rgnum, configday, l34numadd, ip, prefix) + "\n")
                    f_l34_g_add.write(tmpl_l34_g_add % (rgnum, rgnum, configday, l34numadd) + "\n")
                # l34规则是ipv6场景
                elif ipver == 6:
                    # (20210609) 2409:8c20:1831:301:2::/80 中兴UPF返回: 参数IPv6SERVERIP的值 与参数类型定义不一致
                    # 通过 format(netaddr.ipv6_full) 解决
                    ip2 = netaddr.IPAddress(ip).format(netaddr.ipv6_full)
                    f_l34_f_add.write(tmpl_l34_f_v6_add % (rgnum, configday, l34numadd, ip2, prefix) + "\n")
                    f_l34_g_add.write(tmpl_l34_g_add % (rgnum, rgnum, configday, l34numadd) + "\n")
    if len(in_old_l34_name) > 0:
        with open(l34_fl_del, 'w', encoding='utf-8') as f_l34_f_del, \
                open(l34_flgp_del, 'w', encoding='utf-8') as f_l34_g_del:
            for everyold_l34 in in_old_l34_name:
                l34numdel += 1
                f_l34_g_del.write(tmpl_l34_g_del % (rgnum, everyold_l34) + "\n")
                f_l34_f_del.write(tmpl_l34_f_del % everyold_l34 + "\n")
    fndelemptyoutfile(l34_fl_add)
    fndelemptyoutfile(l34_flgp_add)
    fndelemptyoutfile(l34_fl_del)
    fndelemptyoutfile(l34_flgp_del)
    # l7 配置模板
    tmpl_l7_f_add = 'ADD L7FILTER:FILTERNAME="l7_f_%s_%s_%04d",URL="%s",METHOD="METHOD_ANY",APPTYPE="HTTP";'
    tmpl_l7_g_add = 'ADD L7FILTERGROUP:GROUPNAME="l7_g_%s_1",L7FILTERNAME="l7_f_%s_%s_%04d" '
    tmpl_l7_f_del = 'DEL L7FILTER:FILTERNAME="%s" '
    tmpl_l7_g_del = 'DEL L7FILTERGROUP:GROUPNAME="l7_g_%s_1",L7FILTERNAME="%s" '
    # l7 输出配置命令
    l7_fl_add = outdir + '/l7_f_add.txt'
    l7_flgp_add = outdir + '/l7_g_add.txt'
    l7_fl_del = outdir + '/l7_f_del.txt'
    l7_flgp_del = outdir + '/l7_g_del.txt'
    l7numadd = 0
    l7numdel = 0
    if len(in_new_l7) > 0:
        with open(l7_fl_add, 'w', encoding='utf-8') as f_l7_f_add, \
                open(l7_flgp_add, 'w', encoding='utf-8') as f_l7_g_add:
            for every_l7 in in_new_l7:
                l7numadd += 1
                f_l7_f_add.write(tmpl_l7_f_add % (rgnum, configday, l7numadd, every_l7) + "\n")
                f_l7_g_add.write(tmpl_l7_g_add % (rgnum, rgnum, configday, l7numadd) + "\n")
    if len(in_old_l7_name) > 0:
        with open(l7_fl_del, 'w', encoding='utf-8') as f_l7_f_del, \
                open(l7_flgp_del, 'w', encoding='utf-8') as f_l7_g_del:
            for everyold_l7 in in_old_l7_name:
                l7numdel += 1
                f_l7_g_del.write(tmpl_l7_g_del % (rgnum, everyold_l7) + "\n")
                f_l7_f_del.write(tmpl_l7_f_del % everyold_l7 + "\n")
    fndelemptyoutfile(l7_fl_add)
    fndelemptyoutfile(l7_flgp_add)
    fndelemptyoutfile(l7_fl_del)
    fndelemptyoutfile(l7_flgp_del)
    print(u'共生成%d条定向流量规则\n三层共%3d条:新增%3d,删除%3d\n 七层共%3d条:新增%3d,删除%3d' % (
        l34numadd + l34numdel + l7numadd + l7numdel, l34numadd + l34numdel, l34numadd, l34numdel, l7numadd + l7numdel,
        l7numadd, l7numdel))


if __name__ == '__main__':
    upffilterdir = ".\\cur"
    subdir = os.listdir(upffilterdir)
    # 需要更新的定向流量规则
    # jituan_dpifile = u'20121217-小红书.xlsx'
    # jituan_dpifile = u'20121217-爱奇艺.xlsx'
    # jituan_dpifile = u'20121217-手淘天猫.xlsx'
    jituan_dpifile = u'20121217-阿里巴巴.xlsx'
    # jituan_dpifile = u'20121217-阿里云.xlsx'
    time1 = time.time()
    print(str(datetime.now().strftime('%Y-%m-%d %H:%M:%S')), u'开始检测')
    print(str(datetime.now().strftime('%Y-%m-%d %H:%M:%S')), u'开始处理定向流量规则表格')
    print(u'定向流量规则文件:', jituan_dpifile)
    rg_pid, pid_name, new_l34, new_l7 = fn_get_cur_dpi_xlsx(jituan_dpifile)
    # print('-' * 36)
    print(str(datetime.now().strftime('%Y-%m-%d %H:%M:%S')), u'定向流量规则表格读取完毕')
    for everyupf in subdir:
        if not everyupf.startswith("upf"):
            print('skipping subdir', everyupf)
            continue
        upfsubdir = upffilterdir + "\\" + everyupf
        os.chdir(upfsubdir)
        upfl34filer = 'SHOW+L34FILTER_INFO.csv'
        upfl7filter = 'SHOW+L7FILTER_INFO.csv'
        upfl34filergrp = 'SHOW+L34FILTERGROUP_INFO.csv'
        upfl7filtergrp = 'SHOW+L7FILTERGROUP_INFO.csv'
        print('-' * 40)
        print("%s 存量规则处理开始" % everyupf)
        # (20211219) 优化空间，只对需要处理的RG进行加工，非全量RG
        l34filtergrp_dic, l34filtergrp_all, l34filterinfo_all = fn_get_cur_l34filtergrp(upfl34filergrp)
        print(str(datetime.now().strftime('%Y-%m-%d %H:%M:%S')), u'现网UPF L34FILTERGROUP 共', len(l34filterinfo_all),
              '规则组加载完毕')
        l7filtergrp_dic, l7filtergrp_all, l7filterinfo_all = fn_get_cur_l7filtergrp(upfl7filtergrp)
        print(str(datetime.now().strftime('%Y-%m-%d %H:%M:%S')), u'现网UPF  L7FILTERGROUP 共', len(l7filterinfo_all),
              '规则组加载完毕')
        old_l34_name, old_l34 = fn_get_cur_l34filter(upfl34filer, l34filterinfo_all, rg_pid)
        # (DONE) 20210401 7层规则处理
        old_l7_name, old_l7 = fn_get_cur_l7filter(upfl7filter, l7filterinfo_all, rg_pid)
        os.chdir("../../")
        fn_compare_rough(rg_pid, old_l34_name, old_l7_name, new_l34, new_l7, everyupf)
        print("%s 存量规则处理结束" % everyupf)
        # break
    print("程序运行时间 %.3f seconds" % (time.time() - time1))

#! /usr/bin/env python3
# ! -*- coding: utf-8 -*-
# yaozhengyi.ah@chinatelecom Anhui Usage scence
# Compare 5G SA UPF(SHOW L34FILTER; SHOW L34FILTERGROUP; SHOW L7FILTER;SHOW L7FILTERGROUP) to JT rule xlsx file
# (DONE) 20210324 读取UPF存量规则文件作为输入项
# (DONE) 20210329 从 L34/L7FILTERGROUP_INFO 获取到 L34/L7FILTER_INFO 的映射关系
# (DONE) 20210329 对UPF存量规则中特定RG进行L34/L7层规则重复判断 (参数flag_chk_dupl)
# (DONE) 20210402 从xlsx表格中提取需要更新的RG
# (DONE) 20210403 比较xlsx表格与UPF存量规则输出UPF配置脚本
# (DONE) 20210403 根据xlsx表格统计规则时，剔除维护信息为"删除"或“本次删除”
# (DONE) 20210610 适配中兴IPv6地址长度部分识别异常场景。 (中兴仍然存在无法自动转化IPv6前缀的问题!!!,例如 240e:978:1507:0:0:0:0:2/125)
# (DONE) 20210611 对xlsx表格中同时存在删除和非删除的l34或l7规则，以删除为准
# (DONE) 20210707 优化输出UPF配置脚本部分,删除大小为0的配置脚本文件
# (DONE) 20210813 UPF规则差异补齐L34脚本(需结合bcompare软件, grep命令. refer to gendiff_l34.py),差异太多建议全量更新。
# (DONE) 20210827 增加运行过程中输出l7规则新增和删除清单，如输出“请人工复查后再次运行程序”，务必检查xlsx规则文件！！！
# (DONE) 20210906 对N列URL为https/http开头的字符进行剔除,例如(HTTPS https://ecompic0.hitv.com)，剔除https://。 ZTE协议类型没有HTTPS:APPTYPE=HTTP
# (DONE) 20211022 UPF规则差异补齐L7脚本(需结合bcompare软件, grep命令. refer to gendiff_l7.py),差异太多建议全量更新。
# (DONE) 20210801 20211219完成 精确匹配集团规则全量更新 逻辑太简单粗暴了 refer to upfdxllrough.py
# (DONE) 20211219 遍历UPF子目录输出脚本 终于把这个坑填了
# (TODO) 20210801 回退脚本
# 常见RG
# 3532707700    阿里云
# 3532708000    VIVO
# 3532703600    阿里巴巴
# 3532704600    爱奇艺
# 3532707000    百度
# 1600000002    哔哩哔哩
# 3532707600    手淘天猫
# 3532707100    网易
# 3532704700    优酷
# 3532706700    今日头条
# 3532704500    腾讯视频
# 3532706100    天翼云盘

import csv
import os
import re
import time
from datetime import datetime
# 以下两个模块openpyxl和netaddr需要额外安装
import netaddr
import openpyxl
import warnings

warnings.simplefilter("ignore")


# 检查IPv6子网，检查xlsx文件中的子网是否在原UPF规则文件中存在。为避免字符串格式，统一使用IPNetwork来判断
def fnfindmatch(in_every_l34, fnin_old_l34):
    for every_old_l34 in fnin_old_l34:
        if in_every_l34 == netaddr.IPNetwork(every_old_l34):
            return 1
    return 0


# 删除为空的输出文件
def fndelemptyoutfile(infile):
    # (20211224)先判断文件是否存在
    if os.path.exists(infile):
        sz = os.stat(infile).st_size
        if sz == 0:
            os.remove(infile)


# 获得三四层过滤规则组名(L34FILTERGROUP_INFO)与三四层过滤器名(L34FILTER_INFO)的对应关系
def fn_get_cur_l34filtergrp(inl34filtergrp):
    fnl34filtergrp_dic = {}
    fnl34filtergrp_all = []
    fnl34filterinfo_all = []
    with open(inl34filtergrp, 'r', encoding='utf-8') as infile:
        reader = csv.reader(infile)
        # 跳过首行
        next(reader)
        for line in reader:
            l34filtergrp = line[0]
            l34filterinfo = line[1]
            if l34filtergrp in fnl34filtergrp_dic:
                fnl34filtergrp_dic[l34filtergrp].append(l34filterinfo)
            else:
                fnl34filtergrp_all.append(l34filtergrp)
                fnl34filtergrp_dic[l34filtergrp] = [l34filterinfo]
            if l34filterinfo in fnl34filterinfo_all:
                print(u'!!!L34FILTERGROUP_INFO文件中发现重复规则，建议人工核查', l34filterinfo)
            fnl34filterinfo_all.append(l34filterinfo)
    return fnl34filtergrp_dic, fnl34filtergrp_all, fnl34filterinfo_all


# 获得三四层过滤规则组名(L7FILTERGROUP_INFO)与三四层过滤器名(L7FILTER_INFO)的对应关系
def fn_get_cur_l7filtergrp(inl7filtergrp):
    fnl7filtergrp_dic = {}
    fnl7filtergrp_all = []
    fnl7filterinfo_all = []
    with open(inl7filtergrp, 'r', encoding='utf-8') as infile:
        reader = csv.reader(infile)
        # 跳过首行
        next(reader)
        for line in reader:
            l7filtergrp = line[0]
            l7filterinfo = line[1]
            if l7filtergrp in fnl7filtergrp_dic:
                fnl7filtergrp_dic[l7filtergrp].append(l7filterinfo)
            else:
                fnl7filtergrp_all.append(l7filtergrp)
                fnl7filtergrp_dic[l7filtergrp] = [l7filterinfo]
            if l7filterinfo in fnl7filterinfo_all:
                print(u'!!!L7FILTERGROUP_INFO文件中发现重复规则，建议人工核查', l7filterinfo)
            fnl7filterinfo_all.append(l7filterinfo)
    return fnl7filtergrp_dic, fnl7filtergrp_all, fnl7filterinfo_all


def fn_get_rg_from_l34filtergrp(indictkey):
    reobj_rg_name = re.compile(r"_(\d{9,10})")
    match_rg_name = reobj_rg_name.search(indictkey)
    # 得到 过滤规则名称 中的RG值
    if match_rg_name:
        out_rg_name = match_rg_name.group(1)
    else:
        # print(u'L34FILTERGROUP没有找到9或10位长度的RG值', indictkey)
        return None
    return out_rg_name


def fn_get_cur_l34filter(inl34filter_info, in_l34filterinfo_all, chk_rgnum):
    # l34层规则判重开关， 1为开启判断，0为不开启判断
    flag_chk_dupl = 0
    # 输出两个list，分别为规则名称和5元组
    # 实际L34 filtername(过滤规则名称)命名 : l34_f_3532704700_ipv6_20210305_10, l34_f_3532704700_ipv6_169 cry~~~
    # 建议L34 filtername(过滤规则名称) 规范命名 : l34_f_RG_ipv4_YYYYMMDD_NUM/l34_f_RG_ipv6_YYYYMMDD_NUM NUM为四位数字
    # (DONE) 20210330 增加 过滤规则名称 与 过滤规则组的判断
    with open(inl34filter_info, 'r', encoding='utf-8') as infile:
        reader = csv.reader(infile)
        # 跳过首行
        next(reader)
        l34filternamelist = []
        l34filtersubnetlist = []
        duplnum = 0
        for line in reader:
            if "_f_any_" in line or "_f_dns_" in line:
                continue
            reobj_rg_name = re.compile(r"f_(\d{9,10})_")
            match_rg_name = reobj_rg_name.search(line[0])
            # 得到 过滤规则名称 中的RG值
            if match_rg_name:
                fnrg_name = match_rg_name.group(1)
            else:
                continue
                # print(u'L34FILTER', line[0], '没有找到9或10位长度的RG值')
            # 跳过默认规则 RG 3000000000
            if fnrg_name == '3000000000':
                continue
            # 为减少计算量，只对需要更新的RG值获取l34规则
            if not fnrg_name == chk_rgnum:
                continue
            if line[0] not in in_l34filterinfo_all:
                print(u'!!!过滤规则名称(L34FILTER)', line[0], '未找到对应的过滤规则组(L34FILTERGROUP)')
                continue
            if line[1] == 'IPV4':
                subnet = line[2].strip() + "/" + line[3].strip()
            elif line[1] == 'IPV6':
                subnet = line[5].strip() + "/" + line[6].strip()
            l34filternamelist.append(line[0].strip())
            l34filtersubnetlist.append(subnet)
            # l34规则判重开关,打开后增加计算量
            # 判重逻辑：存在网段交集即认为重复
            if flag_chk_dupl == 1:
                if len(l34filtersubnetlist) > 1:
                    for everycursubnet in l34filtersubnetlist[0:-1]:
                        k1 = netaddr.IPNetwork(everycursubnet)
                        k2 = netaddr.IPNetwork(subnet)
                        # (20210414) 只当同为IPv4或者IPv6时进行判重,利还是弊？
                        if k1.version == k2.version:
                            # if netaddr.IPNetwork(subnet) in curipset: //20210328 废弃 速度慢
                            # 优化检测速度，求交集比求子集准确，size比leng() 更快 (实际测试速度从759秒提升到181秒)
                            if (netaddr.IPSet(k2) & netaddr.IPSet(k1)).size > 0:
                                elementindex = l34filtersubnetlist.index(everycursubnet)
                                duplnum += 1
                                print(u'!!!发现l34规则', line[0], subnet, '与l34规则存在重复地址',
                                      l34filternamelist[elementindex], everycursubnet)
                                break
    # 范例数据:
    # ['l34_f_3532704700_ipv4_1', 'l34_f_3532704700_ipv4_10', 'l34_f_3532704700_ipv4_100']
    # ['101.226.184.0/25', '111.13.133.0/24', '183.61.238.0/24', '116.211.221.12/32']
    if duplnum > 0:
        print(u'一共发现%d组重复的l34规则' % duplnum)
    return l34filternamelist, l34filtersubnetlist,


def fn_get_cur_l7filter(inl7filter_info, in_l7filterinfo_all, chk_rgnum):
    # l7层规则判重开关， 1为开启判断，0为不开启判断
    flag_chk_dupl = 1
    # 输出两个list，分别为规则名称,URL
    # 实际L7 filtername(过滤规则名称)命名 : l7_f_1000000001_ipv4_2, l7_f_3532703300_76,l7_f_3532706700_143 cry~~~
    # 建议L7 filtername(过滤规则名称) 规范命名 : l7_f_RG_YYYYMMDD_NUM RG为9-10位数字，YYYYMMDD为日期，NUM为四位数字
    # (DONE) 20210401 增加 过滤规则名称 与 过滤规则组的判断
    # (DONE) 20210401 增加 过滤规则URL重复判断
    # l7规则判重开关,打开后耗费时间较多
    with open(inl7filter_info, 'r', encoding='utf-8') as infile:
        reader = csv.reader(infile)
        # 跳过首行
        next(reader)
        l7filternamelist = []
        l7filterurllist = []
        duplnum = 0
        for line in reader:
            if "filter" in line:
                continue
            reobj_rg_name = re.compile(r"f_(\d{9,10})_")
            match_rg_name = reobj_rg_name.search(line[0])
            if match_rg_name:
                fnrg_name = match_rg_name.group(1)
            else:
                # print(u'L7FILTER', line[0], '没有找到9或10位长度的RG值')
                continue
            # 为减少计算量，只对需要更新的RG值获取l7规则
            if not fnrg_name == chk_rgnum:
                continue
            if line[0] not in in_l7filterinfo_all:
                print(u'!!!过滤规则名称(L7FILTER)', line[0], '未找到对应的过滤规则组(L7FILTERGROUP)')
                continue
            url = line[1].strip()
            l7filternamelist.append(line[0])
            l7filterurllist.append(url)
            if flag_chk_dupl == 1:
                if len(l7filterurllist) > 1:
                    for existurl in l7filterurllist[0:-1]:
                        if url == existurl:
                            existurlindex = l7filterurllist.index(existurl)
                            duplnum += 1
                            print(u'!!!发现l7规则', line[0], url, '与l7规则存在重复URL',
                                  l7filternamelist[existurlindex], existurl)
                            break
    # 范例数据:
    # ['l7_f_3532704700_10', 'l7_f_3532704700_11', 'l7_f_3532704700_12']
    # ['*.cibntv.net', '*.youku.com', '*.uczzd.com']
    if duplnum > 0:
        print(u'一共发现%d组重复的l7规则' % duplnum)
    return l7filternamelist, l7filterurllist


def fn_get_cur_dpi_xlsx(inxlsxfile):
    jtrulewb = openpyxl.load_workbook(inxlsxfile)
    jtrulesheet = jtrulewb[jtrulewb.sheetnames[0]]
    jtrulenumber = jtrulesheet.max_row
    # C3  定向流量RG值
    # I列，IPv4/v6地址
    # J列，掩码
    # N列，URL
    # P列，维护信息，新增/删除/修改/不变/生效中/已报备/本次新增/本次删除 不同规则文件此列填的五花八门，默认按照全量数据比对，额外核对删除部分
    rg_txt = jtrulesheet["C3"].value
    productname = jtrulesheet["A3"].value
    # (DONE) 20211217-20211219 RG查找优化
    # dpirg = re.findall(r'\d+', rg_txt)[-1]
    dpirg = '0'
    digital_all = re.findall(r'\d+(?!\d+)', rg_txt)
    for digital in digital_all:
        if len(digital) == 10:
            dpirg = digital
    if dpirg == '0':
        print(u'!!!定向浏览规则表格文件中的RG长度可能有误，建议人工复查')
    col_i = jtrulesheet["I"]
    # print(col_i[0].value,col_i[1].value,col_i[2].value,col_i[3].value,col_i[-3].value,col_i[-2].value)
    col_j = jtrulesheet["J"]
    col_n = jtrulesheet["N"]
    # print(col_n[0].value, col_n[1].value, col_n[2].value,col_n[3].value,col_n[-3].value，col_n[-2].value)
    col_p = jtrulesheet["P"]
    jtrulewb.close()
    # l34是IPNetwork组成的list
    l34_add = []
    l34_dele = []
    l7_add = []
    l7_dele = []
    for row_num in range(2, jtrulenumber):
        l34ip_value = col_i[row_num].value
        l34mask_value = col_j[row_num].value
        l7host_value = col_n[row_num].value
        flag_value = col_p[row_num].value
        if l34ip_value is None:
            if l7host_value is not None:
                # (DONE) 20210906 对N列URL为https://或http://开头,剔除协议部分  fixed in 20210908
                if l7host_value.startswith('https://'):
                    l7host_value = l7host_value[8::]
                elif l7host_value.startswith('http://'):
                    l7host_value = l7host_value[7::]
                # (20211219) 只有有删除关键字就认为是不需要的
                # if u'删除' in flag_value :
                if flag_value == '删除' or flag_value == '本次删除':
                    l7_dele.append(l7host_value.strip())
                else:
                    l7_add.append(l7host_value.strip())
            else:
                print(u'\033[0;31;40m第', row_num + 1, '行数据的I列目的IP地址无信息，N列URL无信息，请人工复查后再次运行程序!\033[0m')
                break
        elif l34ip_value is not None:
            # （20210401) 检查l34地址是否是有效的地址
            try:
                # I列给出掩码，忽略J列
                if '/' in l34ip_value:
                    l34_vlaue = netaddr.IPNetwork(l34ip_value.strip())
                # I列未给出掩码,取J列作为掩码
                else:
                    l34_vlaue = netaddr.IPNetwork(l34ip_value.strip() + '/' + l34mask_value.strip())
            except Exception as exp:
                print(u'I' + str(row_num) + '单元格内容有误，请人工检查!', l34ip_value.strip(), l34mask_value.strip())
                print(u'错误类型为', exp)
                break
            # if u'删除' in flag_value :
            if l34_vlaue.version == 4:
                if int(l34_vlaue.prefixlen) < 20:
                    print(u'\033[0;31;40mIPv4前缀小于20,请额外注意\033[0m', str(l34_vlaue))
            elif l34_vlaue.version == 6:
                if int(l34_vlaue.prefixlen) <= 64:
                    print(u'\033[0;31;40mIPv6前缀小于63,请额外注意\033[0m', str(l34_vlaue))
            if flag_value == '删除' or flag_value == '本次删除':
                l34_dele.append(l34_vlaue)
                # (20210708)对标记为删除的，额外处理
                if l34_vlaue in l34_add:
                    l34_add.remove(l34_vlaue)
            else:
                if l34_vlaue not in l34_add:
                    l34_add.append(l34_vlaue)
    # 以下对初步得到的l34和l7规则进行去重，删除优先操作
    l34_par1 = []
    # flag: 0 存量 1 增加 -1 删除
    l34_par1_flag = []
    l34_par2 = []
    l34_par2_flag = []
    l7_par1 = []
    l7_par1_flag = []
    l7_par2 = []
    l7_par2_flag = []
    for every_l34_ele in l34_add:
        if every_l34_ele in l34_dele:
            print(every_l34_ele, u'发现在删除清单中')
            continue
        else:
            if every_l34_ele not in l34_par1:
                l34_par1.append(every_l34_ele)
                l34_par1_flag.append(1)
    for every_l34_ele2 in l34_dele:
        if every_l34_ele2 not in l34_par2:
            l34_par2.append(every_l34_ele2)
            l34_par2_flag.append(-1)
    l34_out = l34_par1 + l34_par2
    l34_out_flag = l34_par1_flag + l34_par2_flag
    # 20210827 增加运行过程中输出L7规则
    # print('-'*40)
    # if len(l7_add) > 0:
    #     print(u'l7规则新全量',len(l7_add))
    #     for singlel7add in l7_add:
    #         print(singlel7add)
    # else:
    #     print('无l7规则')
    # print('-'*40)
    # if len(l7_dele) > 0:
    #     print(u'l7规则删除全量',len(l7_dele))
    #     for singlel7del in l7_add:
    #         print(singlel7del)
    # else:
    #     print('无l7规则删除')
    # print('-'*40)
    for every_l7_ele in l7_add:
        if every_l7_ele in l7_dele:
            continue
        else:
            if every_l7_ele not in l7_par1:
                l7_par1.append(every_l7_ele)
                l7_par1_flag.append(1)
    for every_l7_ele2 in l7_dele:
        if every_l7_ele2 not in l7_par2:
            l7_par2.append(every_l7_ele2)
            l7_par2_flag.append(-1)
    l7_out = l7_par1 + l7_par2
    l7_out_flag = l7_par1_flag + l7_par2_flag
    print(u'规则共' + str(jtrulenumber) + u'行')
    print(u'\033[0;30;42m' + productname, 'RG', dpirg + u'\033[0m')
    print(u'三层IP规则(存量/删除) ' + str(len(l34_par1_flag)) + '/' + str(len(l34_par2_flag)) + ' 七层URL规则(存量/删除) ' + str(
        len(l7_par1_flag)) + '/' + str(len(l7_par2_flag)) + u'\033[0m')
    return dpirg, productname, l34_out, l7_out, l34_out_flag, l7_out_flag


def fn_compare(rgnum, in_old_l34_name, in_old_l34, in_old_l7_name, in_old_l7, in_new_l34, in_new_l34_flag,
               in_new_l7, in_new_l7_flag, in_upfnum):
    configday = str(datetime.now().strftime('%Y%m%d'))
    config_time = str(datetime.now().strftime('%Y%m%d%H%M%S'))
    outdir = "rules/" + config_time + "_RG_" + rgnum + "_" + in_upfnum
    # 建议规范写法
    # l34filername:l34_f_{RG}_{YYYYMMDD}_{NUM}
    # l7filername:l7_f_{RG}_{YYYYMMDD}_{NUM}
    # l34filtergrp:l34_g_{RG}_1
    # l7filtergrp:l7_g_{RG}_1
    # RG：9-10位长度的整数
    # YYYYMMDD: 8位长度年月日
    # NUM：四位长度整数
    if not os.path.exists(outdir):
        os.mkdir(outdir)
    # l34 输出配置命令
    l34_fl_add = outdir + '/l34_f_add.txt'
    l34_flgp_add = outdir + '/l34_g_add.txt'
    l34_fl_del = outdir + '/l34_f_del.txt'
    l34_flgp_del = outdir + '/l34_g_del.txt'
    # l34 配置模板
    tmpl_l34_f_v4_add = 'ADD L34FILTER:FILTERNAME="l34_f_%s_%s_%04d",IPTYPE="IPV4",IPV4SERVERIP="%s",' \
                        'IPV4SERVERIPMASK=%s,PROTOCOL="ANY",SERVERPORTSTART=0,SERVERPORTEND=0; '
    tmpl_l34_f_v6_add = 'ADD L34FILTER:FILTERNAME="l34_f_%s_%s_%04d",IPTYPE="IPV6",IPV6SERVERIP="%s",' \
                        'IPV6SERVERIPMASK=%s,PROTOCOL="ANY",SERVERPORTSTART=0,SERVERPORTEND=0;'
    tmpl_l34_f_del = 'DEL L34FILTER:FILTERNAME="%s" '
    tmpl_l34_g_add = 'ADD L34FILTERGROUP:GROUPNAME="l34_g_%s_1",L34FILTERNAME="l34_f_%s_%s_%04d" '
    tmpl_l34_g_del = 'DEL L34FILTERGROUP:GROUPNAME="l34_g_%s_1",L34FILTERNAME="%s" '
    l34numadd = 0
    l34numdel = 0
    if len(in_new_l34) > 0:
        with open(l34_fl_add, 'w', encoding='utf-8') as f_l34_f_add, \
                open(l34_flgp_add, 'w', encoding='utf-8') as f_l34_g_add, \
                open(l34_fl_del, 'w', encoding='utf-8') as f_l34_f_del, \
                open(l34_flgp_del, 'w', encoding='utf-8') as f_l34_g_del:
            for every_l34 in in_new_l34:
                l34index = in_new_l34.index(every_l34)
                # l34规则是ipv4场景
                if every_l34.version == 4:
                    # continue
                    # l34规则为删除
                    if in_new_l34_flag[l34index] == -1:
                        # 旧规则有此条l34
                        if str(every_l34) in in_old_l34:
                            tempindex = in_old_l34.index(str(every_l34))
                            l34numdel += 1
                            f_l34_g_del.write(tmpl_l34_g_del % (rgnum, in_old_l34_name[tempindex]) + "\n")
                            f_l34_f_del.write(tmpl_l34_f_del % (in_old_l34_name[tempindex]) + "\n")
                        # 旧规则无此条l34
                        else:
                            print(u'没有在旧规则中找到需要删除对应的l34规则', str(every_l34))
                    # l34规则为非删除，认为均为存量，与当前的比对
                    else:
                        # 旧规则无此条l34，判断规则为字符型完全匹配。
                        # (TODO) 20211224 后续考虑有交集的场景
                        if str(every_l34) not in in_old_l34:
                            ip, prefix = str(every_l34).split('/')
                            l34numadd += 1
                            f_l34_f_add.write(tmpl_l34_f_v4_add % (rgnum, configday, l34numadd, ip, prefix) + "\n")
                            f_l34_g_add.write(tmpl_l34_g_add % (rgnum, rgnum, configday, l34numadd) + "\n")
                elif every_l34.version == 6:
                    # l34规则为删除
                    if in_new_l34_flag[l34index] == -1:
                        # 旧规则有此条l34
                        if str(every_l34) in in_old_l34:
                            tempindex = in_old_l34.index(str(every_l34))
                            l34numdel += 1
                            f_l34_g_del.write(tmpl_l34_g_del % (rgnum, in_old_l34_name[tempindex]) + "\n")
                            f_l34_f_del.write(tmpl_l34_f_del % (in_old_l34_name[tempindex]) + "\n")
                        # 旧规则无此条l34
                        else:
                            print(u'没有在旧规则中找到需要删除对应的l34规则', str(every_l34))
                    # l34规则为非删除，认为均为存量，与当前的比对
                    else:
                        # 旧规则无此条l34
                        if fnfindmatch(every_l34, in_old_l34) == 0:
                            ip, prefix = str(every_l34).split('/')
                            l34numadd += 1
                            # (20210609) 2409:8c20:1831:301:2::/80 中兴UPF返回: 参数IPv6SERVERIP的值 与参数类型定义不一致
                            # 通过 format(netaddr.ipv6_full) 解决
                            ip2 = netaddr.IPAddress(ip).format(netaddr.ipv6_full)
                            f_l34_f_add.write(
                                tmpl_l34_f_v6_add % (rgnum, configday, l34numadd, ip2, prefix) + "\n")
                            f_l34_g_add.write(tmpl_l34_g_add % (rgnum, rgnum, configday, l34numadd) + "\n")
        fndelemptyoutfile(l34_fl_add)
        fndelemptyoutfile(l34_flgp_add)
        fndelemptyoutfile(l34_fl_del)
        fndelemptyoutfile(l34_flgp_del)
    # l7 配置模板
    tmpl_l7_f_add = 'ADD L7FILTER:FILTERNAME="l7_f_%s_%s_%04d",URL="%s",METHOD="METHOD_ANY",APPTYPE="HTTP";'
    tmpl_l7_g_add = 'ADD L7FILTERGROUP:GROUPNAME="l7_g_%s_1",L7FILTERNAME="l7_f_%s_%s_%04d" '
    tmpl_l7_f_del = 'DEL L7FILTER:FILTERNAME="%s" '
    tmpl_l7_g_del = 'DEL L7FILTERGROUP:GROUPNAME="l7_g_%s_1",L7FILTERNAME="%s" '
    # l7 输出配置命令
    l7_fl_add = outdir + '/l7_f_add.txt'
    l7_flgp_add = outdir + '/l7_g_add.txt'
    l7_fl_del = outdir + '/l7_f_del.txt'
    l7_flgp_del = outdir + '/l7_g_del.txt'
    l7numadd = 0
    l7numdel = 0
    if len(in_new_l7) > 0:
        with open(l7_fl_add, 'w', encoding='utf-8') as f_l7_f_add, \
                open(l7_flgp_add, 'w', encoding='utf-8') as f_l7_g_add, \
                open(l7_fl_del, 'w', encoding='utf-8') as f_l7_f_del, \
                open(l7_flgp_del, 'w', encoding='utf-8') as f_l7_g_del:
            for every_l7 in in_new_l7:
                l7index = in_new_l7.index(every_l7)
                # l7规则为删除l
                if in_new_l7_flag[l7index] == -1:
                    # 旧规则有此条l7
                    if every_l7 in in_old_l7:
                        tempindex = in_old_l7.index(every_l7)
                        l7numdel += 1
                        f_l7_g_del.write(tmpl_l7_g_del % (rgnum, in_old_l7_name[tempindex]) + "\n")
                        f_l7_f_del.write(tmpl_l7_f_del % (in_old_l7_name[tempindex]) + "\n")
                    # 旧规则无此条l7
                    else:
                        print(u'没有在旧规则中找到需要删除对应的l7规则', str(every_l7))
                # l7规则为非删除，认为均为存量，与当前的比对
                else:
                    # 旧规则无此条l7
                    if str(every_l7) not in in_old_l7:
                        l7numadd += 1
                        f_l7_f_add.write(tmpl_l7_f_add % (rgnum, configday, l7numadd, every_l7) + "\n")
                        f_l7_g_add.write(tmpl_l7_g_add % (rgnum, rgnum, configday, l7numadd) + "\n")
        fndelemptyoutfile(l7_fl_add)
        fndelemptyoutfile(l7_flgp_add)
        fndelemptyoutfile(l7_fl_del)
        fndelemptyoutfile(l7_flgp_del)
    print(u'共生成%d条定向流量规则\n三层共%3d条:新增%3d,删除%3d\n 七层共%3d条:新增%3d,删除%3d' % (
        l34numadd + l34numdel + l7numadd + l7numdel, l34numadd + l34numdel, l34numadd, l34numdel, l7numadd + l7numdel,
        l7numadd, l7numdel))


if __name__ == '__main__':
    upffilterdir = ".\\cur"
    subdir = os.listdir(upffilterdir)
    # 需要更新的定向流量规则
    # jituan_dpifile = u'20121217-小红书.xlsx'
    # jituan_dpifile = u'20121217-爱奇艺.xlsx'
    # jituan_dpifile = u'20121217-手淘天猫.xlsx'
    jituan_dpifile = u'20121217-阿里云.xlsx'
    # jituan_dpifile = u'20121217-阿里巴巴.xlsx'
    time1 = time.time()
    print(str(datetime.now().strftime('%Y-%m-%d %H:%M:%S')), u'开始检测')
    print(str(datetime.now().strftime('%Y-%m-%d %H:%M:%S')), u'开始处理定向流量规则表格')
    print(u'定向流量规则文件:', jituan_dpifile)
    rg_pid, pid_name, new_l34, new_l7, new_l34_flag, new_l7_flag = fn_get_cur_dpi_xlsx(jituan_dpifile)
    # print('-' * 40)
    print(str(datetime.now().strftime('%Y-%m-%d %H:%M:%S')), u'定向流量规则表格读取完毕')
    for everyupf in subdir:
        if not everyupf.startswith("upf"):
            print('skipping subdir', everyupf)
            continue
        upfsubdir = upffilterdir + "\\" + everyupf
        os.chdir(upfsubdir)
        upfl34filer = 'SHOW+L34FILTER_INFO.csv'
        upfl7filter = 'SHOW+L7FILTER_INFO.csv'
        upfl34filergrp = 'SHOW+L34FILTERGROUP_INFO.csv'
        upfl7filtergrp = 'SHOW+L7FILTERGROUP_INFO.csv'
        print('-' * 40)
        print(str(datetime.now().strftime('%Y-%m-%d %H:%M:%S')), everyupf, u'存量规则处理开始')
        l34filtergrp_dic, l34filtergrp_all, l34filterinfo_all = fn_get_cur_l34filtergrp(upfl34filergrp)
        print(str(datetime.now().strftime('%Y-%m-%d %H:%M:%S')), u'现网UPF L34FILTERGROUP 共', len(l34filterinfo_all),
              '规则组加载完毕')
        l7filtergrp_dic, l7filtergrp_all, l7filterinfo_all = fn_get_cur_l7filtergrp(upfl7filtergrp)
        print(str(datetime.now().strftime('%Y-%m-%d %H:%M:%S')), u'现网UPF  L7FILTERGROUP 共', len(l7filterinfo_all),
              '规则组加载完毕')
        old_l34_name, old_l34 = fn_get_cur_l34filter(upfl34filer, l34filterinfo_all, rg_pid)
        # (DONE) 20210401 7层规则处理
        old_l7_name, old_l7 = fn_get_cur_l7filter(upfl7filter, l7filterinfo_all, rg_pid)
        # (DONE) 20210403 根据规则xlsx文件统计规则数量时，剔除维护信息为"删除"
        print(u'\033[0;30;42m' + pid_name, 'RG', rg_pid, ':', '三层IP规则(旧/新) ' + str(len(old_l34)) + '/' + str(
            len(new_l34) - new_l34_flag.count(-1)) + ' 七层URL规则(旧/新) ' + str(
            len(old_l7)) + '/' + str(len(new_l7) - new_l7_flag.count(-1)) + u'\033[0m')
        os.chdir("../../")
        fn_compare(rg_pid, old_l34_name, old_l34, old_l7_name, old_l7,
                   new_l34, new_l34_flag, new_l7, new_l7_flag, everyupf)
        print(str(datetime.now().strftime('%Y-%m-%d %H:%M:%S')), everyupf, u'存量规则处理完毕')
        # break
    print("程序运行时间 %.3f seconds" % (time.time() - time1))
